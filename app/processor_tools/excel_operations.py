from openpyxl.worksheet import worksheet
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
from typing import Dict
from statement_analyser_personal.logger import get_logger

logger = get_logger(__name__)

class ExcelManager():
    def __init__(self, bank_name:str, date:Dict[str,str], results:Dict[str, Dict[str,float]]):
        self.bank_name = bank_name
        self.date = date
        self.results = results      
        logger.info(f"excel_operations_init: ExcelManager initialized for {self.bank_name}. date and results obtained")
        

    def set_alignment(self, ws:worksheet, record_no:int, result: Dict[str, Dict[str, float]] ):
        try:
            logger.info(f"set_alignment: Setting alignment for excel")
            for row in range(2+(record_no-1)*9, 10+(record_no-1)*9):
                for col in range(1, 5+len(result) ):
                    cell = ws.cell(row = row, column = col)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            logger.info(f"set_alignment: Alignment set successfully")
        except Exception as e:
            logger.error(f"set_alignment: Failed to set alignment: {e}")

    def insert_header(self, ws:worksheet, record_no:int, data: Dict[str, str]):
        logger.info(f"insert_header: Writing header to excel")
        list = ["Card Name","Card No.", "Previous Balance", "Credit Payment", "Debit Fees", "Retail Purchases", "Balance Due", "Minimum Payment"]
        ws[f"A{2 + (record_no -1)*9}"] = "Record No."
        ws[f"A{3 + (record_no -1)*9}"] = data["record_number"]
        ws[f"D{3 + (record_no -1)*9}"] = "Total"
        ws[f"B{2 + (record_no -1)*9}"] = "Statement Date"
        ws[f"B{3 + (record_no -1)*9}"] = data["statement_date"]
        ws[f"B{8 + (record_no -1)*9}"] = "Payment Due Date"
        ws[f"B{9 + (record_no -1)*9}"] = data["payment_date"]
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 17
        ws.column_dimensions["C"].width = 17
        ws.column_dimensions["D"].width = 12
        for i, header in enumerate(list, start = 1):
            i += 1+(record_no-1)*9
            ws[f"C{i}"] = header
        logger.info(f"insert_header: Headers written successfully")

    def insert_key_value(self, ws:worksheet, record_no:int, result: Dict[str, Dict[str,float]]):
        logger.info(f"insert_key_value: Writing key value to excel")
        total_dict = {
            "previous_balance": 0,
            "credit_payment": 0,
            "debit_fees": 0,
            "retail_purchase": 0,
            "balance_due": 0,
            "minimum_payment": 0
        }
        for i, (card, values) in enumerate(result.items(), start = 1):
                i += 4
                ws.cell(row= 2+(record_no-1)*9, column = i, value = values["card_name"]),
                ws.cell(row= 2+(record_no-1)*9, column = i).font = Font(size = 9.5, bold = True)
                ws.cell(row= 3+(record_no-1)*9, column = i, value = card)
                ws.cell(row= 3+(record_no-1)*9, column = i).font = Font(bold = True)
                ws.cell(row= 4+(record_no-1)*9, column = i, value = values["previous_balance"]),
                ws.cell(row= 5+(record_no-1)*9, column = i, value = values["credit_payment"]),
                ws.cell(row= 6+(record_no-1)*9, column = i, value = values["debit_fees"]),
                ws.cell(row= 7+(record_no-1)*9, column = i, value = values["retail_purchase"]),    
                ws.cell(row= 8+(record_no-1)*9, column = i, value = values["balance_due"]),
                ws.cell(row= 9+(record_no-1)*9, column = i, value = values["minimum_payment"]),
                logger.info(f"insert_key_value: Writing values of {card} to excel")
                total_dict["previous_balance"] += values["previous_balance"]
                total_dict["credit_payment"] += values["credit_payment"]
                total_dict["debit_fees"] += values["debit_fees"]
                total_dict["retail_purchase"] += values["retail_purchase"]
                total_dict["balance_due"] += values["balance_due"]
                total_dict["minimum_payment"] += values["minimum_payment"]
                logger.info(f"insert_key_value: Adding amount to total dict")
                ws.column_dimensions[get_column_letter(i)].width = 20
        ws.row_dimensions[2+(record_no-1)*9].height = 26
        logger.info(f"insert_key_value: Finished writing key value to excel")
        logger.info(f"insert_key_value: Writing total values to excel")   
        ws.cell(row= 4+(record_no-1)*9, column = 4, value = total_dict["previous_balance"]),
        ws.cell(row= 5+(record_no-1)*9, column = 4, value = total_dict["credit_payment"]),
        ws.cell(row= 6+(record_no-1)*9, column = 4, value = total_dict["debit_fees"]),
        ws.cell(row= 7+(record_no-1)*9, column = 4, value = total_dict["retail_purchase"]),    
        ws.cell(row= 8+(record_no-1)*9, column = 4, value = total_dict["balance_due"]),
        ws.cell(row= 9+(record_no-1)*9, column = 4, value = total_dict["minimum_payment"]),
        logger.info(f"insert_key_value: Finished writing total values to excel")
        

    def create_excel_file(self,save_path = None):
        logger.info(f"create_excel_file: Creating new Excel file")

        data = {
            "record_number" : 1,
            "statement_date" : self.date["statement_date"],
            "payment_date" : self.date["payment_date"],
            "bank_name" : self.bank_name,
            "results1" : self.results
        }

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = self.bank_name
            logger.info(f"create_excel_file: Worksheet {self.bank_name} created")

            self.insert_everything(ws, 1, data)
            
            if not save_path:
                logger.error("create_excel_file: No save_path provided for create_excel_file in GUI mode.")
                raise ValueError("create_excel_file: No save_path provided for create_excel_file in GUI mode.")
            wb.save(save_path)

            logger.info(f"create_excel_file: Saving excel file")

            self.create_update_total_sheet(save_path)

            

            logger.info(f"create_excel_file: Excel file saved at: {save_path}")

        except Exception as e:
            logger.error(f"create_excel_file: Failed to create Excel file: {e}")
            raise RuntimeError(f"create_excel_file: Failed to create Excel file: {str(e)}")

    

    def update_excel(self, excel_path:str):
        logger.info(f"update_excel: Updating existing Excel file at {excel_path}")
        try:
            wb = load_workbook(excel_path)
            logger.info(f"update_excel: Available worksheets: {wb.sheetnames}")

            if self.bank_name in wb.sheetnames:
                ws = wb[self.bank_name]
                record_no = self.find_largest_record_no(ws) +1
                data = {
                        "record_number" : record_no,
                        "statement_date" : self.date["statement_date"],
                        "payment_date" : self.date["payment_date"],
                        "bank_name" : self.bank_name,
                        f"results{record_no}" : self.results
                    }
                logger.info(f"update_excel: Sheet '{self.bank_name}' found in the workbook.")

                self.insert_everything(ws, record_no, data)
                wb.save(excel_path)
                logger.info(f"update_excel: Saving excel file")

                self.create_update_total_sheet(excel_path)

                
                logger.info(f"update_excel: Excel file saved and updated successdully")
                return "updated"
            else:
                logger.info("update_excel: No bank sheet detected in available sheets, creating new one")
                ws = wb.create_sheet(self.bank_name)

                data = {
                        "record_number" : 1,
                        "statement_date" : self.date["statement_date"],
                        "payment_date" : self.date["payment_date"],
                        "bank_name" : self.bank_name,
                        f"results1" : self.results
                    }

                logger.info(f"update_excel: New sheet '{self.bank_name}' created.")
                self.insert_everything(ws, 1, data)
                
                wb.save(excel_path)
                logger.info(f"update_excel: Saving excel file")

                self.create_update_total_sheet(excel_path)

                

                logger.info(f"update_excel: New sheet created and saved successfully.")
                return "inserted"
        except Exception as e:
            logger.error(f"update_excel: Failed to update Excel file: {e}")
            raise RuntimeError(f"update_excel: Failed to update Excel file: {str(e)}")

    def insert_everything(self, ws: worksheet, record_no:int, data: Dict[str,str]):
        
        try:
            logger.info("insert_everything: Writing everythin to excel")
            self.insert_header(ws, record_no, data)

            result = data[f"results{record_no}"]
            self.insert_key_value(ws, record_no, result)

            self.set_alignment(ws, record_no, result)
            
            logger.info(f"insert_everything: Finished writing everything to excel.")
        except Exception as e:
            logger.error(f"insert_everything: Failed to insert info: {e}")
            raise RuntimeError(f"insert_everything: Failed to create Excel file: {str(e)}")

    def find_largest_record_no(self, ws:worksheet):
        try:
            logger.info(f"find_largest_record_no: Finding most recent record number")
            logger.info(f"find_largest_record_no: Loading sheet {ws.title}")
            max_record = None
            empty_count = 0
            for row in range(3, ws.max_row+1):
                cell = ws.cell(row = row, column = 1)
                value = cell.value
                if value is None or isinstance(value, str):
                    empty_count += 1
                    if empty_count > 10:
                        break
                    continue
                empty_count = 0
                try:
                    record_no = int(value)
                    if max_record is None or record_no > max_record:
                        max_record = record_no
                except (TypeError, ValueError):
                    continue
            logger.info(f"find_largest_record_no: Found recent record number: {max_record} for {ws.title}")
            return max_record if max_record is not None else 1
        except Exception as e:
            logger.error(f"find_largest_record_no: Failed to find morst recent record no")
            raise RuntimeError(f"find_largest_record_no: Failed to find most recent record no: {str(e)}")

    def create_update_total_sheet(self, excel_path:str):
        wb = load_workbook(excel_path)
        logger.info(f"create_update_total_sheet: Loading workbook")
        try:
            if "Total" not in wb.sheetnames:
                logger.info(f"create_update_total_sheet: Total sheet not found, creating total sheet")
                ws_total = wb.create_sheet("Total")
            else:
                ws_total = wb["Total"]
                logger.info(f"create_update_total_sheet: Total sheet detected, loading it")
            headers = ["Bank", "Balance Due", "Minimum Payment", "Payment Due Date"]
            for i, header in enumerate(headers, start = 1):
                ws_total.cell(row = 1, column = i, value = header)
                ws_total.cell(row = 1, column = i ).font = Font(bold = True)
            ws_total.column_dimensions["A"].width = 17
            ws_total.column_dimensions["B"].width = 15
            ws_total.column_dimensions["C"].width = 15
            ws_total.column_dimensions["D"].width = 20
            logger.info(f"create_update_total_sheet: Loading all sheets except total")
            sheets = [sheet for sheet in wb.sheetnames if sheet != "Total"]
            logger.info(f"create_update_total_sheet: Available sheets: {sheets}")
            data = {}
            for sheet in sheets:
                ws = wb[sheet]
                
                record_no = self.find_largest_record_no(ws)
                logger.info(f"create_update_total_sheet: Getting data for {sheet}")
                data[sheet] = {}
                cell1 = ws[f'd{8+(record_no-1)*9}']
                cell2 = ws[f'd{9+(record_no-1)*9}']
                cell3 = ws[f'B{9+(record_no-1)*9}']
                logger.info(f"Checking all cell values whether is formula/none/integer")
                def get_cell_value(cell):
                    # If value is None, try to get the formula string
                    if cell.data_type == 'f':
                        if cell.value is None:
                            return cell._value  # This is the formula string, e.g. '=SUM(...)'
                        return cell.value  # This is the cached value if available
                    return cell.value if cell.value is not None else 0

                data[sheet]["Balance Due"] = get_cell_value(cell1)
                logger.info(f"create_update_total_sheet: Balance Due: {data[sheet]['Balance Due']}")
                data[sheet]["Minimum Payment"] = get_cell_value(cell2)
                logger.info(f"create_update_total_sheet: Minimum Payment: {data[sheet]['Minimum Payment']}")
                data[sheet]["Payment Due Date"] = get_cell_value(cell3)
                logger.info(f"create_update_total_sheet: Payment Due Date: {data[sheet]['Payment Due Date']}")
            
            logger.info(f"create_update_total_sheet: Finished getting data from all sheets\n{data}\now writing to total sheet")
            for row, (sheet, values) in enumerate(data.items(), start = 2):
                ws_total.cell(row = row, column = 1, value = sheet)
                ws_total.cell(row = row, column = 1).font = Font(bold = True)
                ws_total.cell(row = row, column = 2, value = values["Balance Due"])
                ws_total.cell(row = row, column = 3, value = values["Minimum Payment"])
                ws_total.cell(row = row, column = 4, value = values["Payment Due Date"])
                logger.info(f"create_update_total_sheet: Finished writing data of {sheet} to total sheet")
            logger.info("create_update_total_sheet: Finished writing all data to total sheet")
            logger.info("create_update_total_sheet: Setting alignment for total sheet")
            for row in range(1,len(data)+2):
                for col in range(1, 5):
                    ws_total.cell(row= row, column = col).alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
            logger.info("create_update_total_sheet: Finished setting alignment for total sheet")
            logger.info("create_update_total_sheet: Finished creating/updating total sheet")
            wb.save(excel_path)
        except Exception as e:
            logger.error(f"create_update_total_sheet: Error: {e}")
            raise RuntimeError(f"create_update_total_sheet: Error: {str(e)}")
            



