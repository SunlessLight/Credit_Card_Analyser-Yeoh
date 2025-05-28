from credit_card_tracker.app.processor_tools.excel_operations import ExcelManager
from credit_card_tracker.app.banks.test import TestUOB, TestHLB, TestCIMB, TestMYB, TestRHB, TestPBB
import os
import json
from typing import Dict
from openpyxl import Workbook
from credit_card_tracker.logger import get_logger
logger = get_logger(__name__)

result = TestUOB().test_process_block()

class TestExcelManager(TestUOB):
    def __init__(self):
        
        self.bank = "UOB"
        self.date = {"statement_date": "2023-10-01", "payment_date": "2023-10-15"}
        self.result = result

    def create_json_file(self, date: Dict[str,str], bank_name: str, results: Dict[str, Dict[str,float]]):
        data = {
            "record_number" : 1,
            "statement_date" : date["statement_date"],
            "payment_date" : date["payment_date"],
            "bank_name" : bank_name,
            "results1" : results
        }
        try:
            data_folder = os.path.join(os.path.dirname(__file__), "..",".." "data")
            os.makedirs(data_folder, exist_ok=True)
            json_path = os.path.join(data_folder, f"{bank_name}.json")
            with open(json_path, "w") as f:
                json.dump(data, f, indent = 2)
                logger.info(f"JSON file created successfully at {bank_name}.json")
            
        except Exception as e:
            logger.error(f"Failed to create JSON file: {e}")
            raise RuntimeError(f"Failed to create JSON file: {str(e)}")
        
    def create_excel_file(self, bank_name : str):
        import tkinter as tk
        from tkinter import filedialog
        data_folder = os.path.join(os.path.dirname(__file__), "..",".." "data")
        json_path = os.path.join(data_folder, f"{bank_name}.json")
        with open(json_path, "r") as f:
            data = json.load(f)
        try:
            
            wb = Workbook()
            ws = wb.active
            ws.title = bank_name
            list = ["Card No.", "Previous Balance", "Credit Payment", "Debit Fees", "Retail Purchases", "Balance Due", "Minimum Payment"]
            ws["A3"] = "Record No."
            ws["A4"] = data["record_number"]
            ws["B3"] = "Statement Date"
            ws["B4"] = data["statement_date"]
            ws["B8"] = "Payment Due Date"
            ws["B9"] = data["payment_date"]
            for i, header in enumerate(list, start = 1):
                i += 2
                ws[f"C{i}"] = header
            result = data["results1"]
            for i, (card, values) in enumerate(result.items(), start = 1):
                i += 3
                ws.cell(row= 1, column = i, value = bank_name),
                ws.cell(row= 3, column = i, value = card),
                ws.cell(row= 4, column = i, value = values["previous_balance"]),
                ws.cell(row= 5, column = i, value = values["credit_payment"]),
                ws.cell(row= 6, column = i, value = values["debit_fees"]),
                ws.cell(row= 7, column = i, value = values["retail_purchase"]),    
                ws.cell(row= 8, column = i, value = values["balance_due"]),
                ws.cell(row= 9, column = i, value = values["minimum_payment"]),
        
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"{bank_name}.xlsx",
                title="Save Excel File As"
            )
            root.destroy()

            if save_path:
                wb.save(save_path)
                print(f"Excel file saved at: {save_path}")
            else:
                print("Save cancelled by user.")
        except Exception as e:
            logger.error(f"Failed to create Excel file: {e}")
            raise RuntimeError(f"Failed to create Excel file: {str(e)}")


test = TestExcelManager()
test.create_json_file(test.date, test.bank, test.result)
test.create_excel_file(test.bank)