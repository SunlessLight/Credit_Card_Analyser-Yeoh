from pathlib import Path
import fitz
from typing import Dict, List, Optional
from banks import UOB, HLB, MYB, RHB, PBB, CIMB
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter,column_index_from_string
import json
from copy import copy
from openpyxl.worksheet.worksheet import Worksheet

CARD_ORDERED_MAP_PATH = Path(__file__).parent / "card_order_map.json"


if not CARD_ORDERED_MAP_PATH.exists():
    with open(CARD_ORDERED_MAP_PATH, "w", encoding="utf-8") as f:
        json.dump({}, f, indent=2)

class CreditCardProcessor:
    BANK_CLASSES = {
        "UOB": UOB,
        "HLB": HLB,
        "MYB": MYB,
        "RHB": RHB,
        "PBB": PBB,
        "CIMB": CIMB,
    }

    
    
    
    def __init__(self, bank: str):
        self.bank = self.BANK_CLASSES[bank]()
        self.bank_name = bank
        if CARD_ORDERED_MAP_PATH.exists():
            print("json file located and used")
            with open(CARD_ORDERED_MAP_PATH, "r", encoding="utf-8") as f:
                self.CARD_ORDERED_MAP = json.load(f)

        else:
            print("json file missing")
        self.card_global_order = self._generate_ordered_list()
        self.card_column_map = {
            suffix: col for suffix, col, _ in self.card_global_order
        }
        print(self.CARD_ORDERED_MAP)  #debug
        print(self.card_global_order)  #debug
        print(self.card_column_map)  #debug

        
        

    def _generate_ordered_list(self) -> List[tuple]:
        return [
            (suffix, col, bank)
            for bank, cards in self.CARD_ORDERED_MAP.items()
            for suffix, col in cards
        ]

    def insert_column_with_format_and_merge(self,ws: Worksheet, idx: int):
        ws.insert_cols(idx)

        max_row = ws.max_row

        # Preserve styles and formatting
        for row in range(1, max_row + 1):
            old_cell = ws.cell(row=row, column=idx + 1)
            new_cell = ws.cell(row=row, column=idx)
            if old_cell.has_style:
                new_cell._style = copy(old_cell._style)
            new_cell.font = copy(old_cell.font)
            new_cell.border = copy(old_cell.border)
            new_cell.fill = copy(old_cell.fill)
            new_cell.number_format = copy(old_cell.number_format)
            new_cell.protection = copy(old_cell.protection)
            new_cell.alignment = copy(old_cell.alignment)

        # Copy column width
        col_letter_old = get_column_letter(idx + 1)
        col_letter_new = get_column_letter(idx)
        if col_letter_old in ws.column_dimensions:
            ws.column_dimensions[col_letter_new].width = ws.column_dimensions[col_letter_old].width

        # Handle merged cells
        updated_merges = []
        to_remove = []

        for merged in ws.merged_cells.ranges:
            min_col, max_col = merged.min_col, merged.max_col

            # Insertion point is before the merged range
            if max_col < idx:
                updated_merges.append(merged)
            # Insertion point is after the merged range
            elif min_col > idx:
                merged.shift(0, 1)
                updated_merges.append(merged)
            # Insertion point is inside the merged range – must split it
            else:
                to_remove.append(merged)

        # Remove conflicting merges
        for r in to_remove:
            ws.unmerge_cells(range_string=str(r))

        # Apply adjusted merges
        ws.merged_cells.ranges = []
        for r in updated_merges:
            ws.merged_cells.add(r)

    def insert_new_card_column(self, new_card: str, after_card: str, excel_path: str):
        # Load the workbook and worksheet
        wb = load_workbook(excel_path)
        ws = wb.active

        # Find the index of the after_card in the global order
        idx = next((i for i, (suffix, _, _) in enumerate(self.card_global_order) if suffix == after_card), None)
        if idx is None:
            raise ValueError(f"After card {after_card} not found in global order.")

        # Get the column letter and index of the after_card
        after_col_letter = self.card_global_order[idx][1]
        after_col_index = column_index_from_string(after_col_letter)
        print(after_col_letter)

        # Insert a new column in the Excel sheet
        self.insert_column_with_format_and_merge(ws, after_col_index + 1)
        
        print("inserted new column")

        # Save the updated Excel file
        wb.save(excel_path)
        print(f"✔ Inserted new column for card {new_card} after {after_card} in Excel.")
        
    
    
        # Calculate the new column letter
        new_col_index = after_col_index + 1
        new_col_letter = get_column_letter(new_col_index)

        # Update the global order with the new card
        self.card_global_order.insert(idx + 1, (new_card, new_col_letter, self.bank_name))

        # Update the column letters for all subsequent cards in the global order
        for i in range(idx + 2, len(self.card_global_order)):
            old_suffix, old_col, old_bank = self.card_global_order[i]
            old_index = column_index_from_string(old_col)
            self.card_global_order[i] = (old_suffix, get_column_letter(old_index + 1), old_bank)
            print(self.card_global_order[i])  #debug

        # Update the card column map
        self.card_column_map = {
            suffix: col for suffix, col, _ in self.card_global_order
        }
        

        print(f"Inserted new column: {new_card} → {new_col_letter}")
        print("Updated card_column_map:")
        for k, v in self.card_column_map.items():
            print(f"  {k}: {v}")
        
        from collections import defaultdict

        new_card_map = defaultdict(list)
        for suffix, col, bank in self.card_global_order:
            new_card_map[bank].append([suffix, col])

        # Replace and save
        self.CARD_ORDERED_MAP = dict(new_card_map)

        self.save_card_order_map()   


    def extract_text(self, pdf_path: str, password: Optional[str] = None) -> List[str]:
        """Extract text from PDF and automatically save raw text file"""
        try:
            doc = fitz.open(pdf_path)
            if doc.is_encrypted:
                if not password or not doc.authenticate(password):
                    raise ValueError("Failed to decrypt PDF")
            
            lines = [
                line.strip() for page in doc
                for line in page.get_text("text").split("\n")
                if line.strip()
            ]
            
            # Always save raw text
            text_path = str(Path(pdf_path).with_suffix('.raw.txt'))
            with open(text_path, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))
            print(f"Saved raw text to: {text_path}")
            
            return lines
            
        except Exception as e:
            raise RuntimeError(f"PDF extraction failed: {str(e)}")

    def parse_statement(self, pdf_path: str, password: Optional[str] = None) -> Dict[str, Dict[str, float]]:
        """Always generates both raw text and blocks files"""
        lines = self.extract_text(pdf_path, password)  # extract_text now saves raw text
        blocks = self.bank.create_blocks(lines)
        
        # Always save blocks
        blocks_path = str(Path(pdf_path).with_suffix('.blocks.txt'))
        with open(blocks_path, "w", encoding="utf-8") as f:
            for card, block_lines in blocks.items():
                f.write(f"===== Card: {card} =====\n")
                f.write("\n".join(block_lines) + "\n\n")
        print(f"Saved blocks to: {blocks_path}")
        
        if hasattr(self.bank, "extract"):
            
            return self.bank.extract(lines)
        else:
            
            return {
        
                card: self.bank.process_block(block, lines)
                for card, block in blocks.items()
            }   
        
    def write_to_excel(self, results: Dict[str, Dict[str, float]], excel_path: str, record_number: int):
        try:
            wb = load_workbook(excel_path)
            ws = wb.active

            base_row = 4 + (record_number - 1) * 7  # Row 4 is start of record 1
            bank_cards = dict(self.CARD_ORDERED_MAP[self.bank_name])
            print(bank_cards)
            for card, data in results.items():
                card_col = bank_cards.get(card)
                if not card_col:
                    print(f"Card {card} not found. Adding to sheet.")
                    bank_cards = [sfx for sfx, _, bk in self.card_global_order if bk == self.bank_name]
                    if bank_cards:
                        self.insert_new_card_column(card, bank_cards[-1], excel_path)
                        wb = load_workbook(excel_path)
                        ws = wb.active
                        bank_cards = dict(self.CARD_ORDERED_MAP[self.bank_name])  # update after insert
                        card_col = bank_cards.get(card)
                    else:
                        raise ValueError(f"No base card found for bank {self.bank_name} to insert {card}.")
                    ws[f"{card_col}{3}"] = card

                # Fill each value in the right row
                ws[f"{card_col}{base_row + 0}"] = data["previous_balance"]
                ws[f"{card_col}{base_row + 1}"] = data["credit_payment"]
                ws[f"{card_col}{base_row + 2}"] = data["debit_fees"]
                ws[f"{card_col}{base_row + 3}"] = data["retail_purchases"]
                ws[f"{card_col}{base_row + 4}"] = data["balance_due"]
                ws[f"{card_col}{base_row + 5}"] = data["minimum_payment"]

                print(f"✔ Updated card {card} in column {card_col}")

            wb.save(excel_path)
            print(f"✅ Excel updated and saved: {excel_path}")

        except Exception as e:
            raise RuntimeError(f"Failed to write to Excel: {str(e)}")
        
    def save_card_order_map(self):
        try:
            with open(CARD_ORDERED_MAP_PATH, "w", encoding="utf-8") as f:
                json.dump(self.CARD_ORDERED_MAP, f, indent=2)
            print(f"✅ Saved card order map to {CARD_ORDERED_MAP_PATH}")
        except Exception as e:
            print(f"Failed to save card order map: {e}")

        
    