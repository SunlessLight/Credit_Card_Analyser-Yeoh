from pathlib import Path
import fitz
from typing import Dict, List, Optional
from credit_card_parser.banks import UOB, HLB, MYB, RHB, PBB, CIMB
from openpyxl import load_workbook
from excel_operations import ExcelManager
from text_extractor import TextExtractor
from openpyxl.utils import get_column_letter, column_index_from_string 



class CreditCardProcessor:
    BANK_CLASSES = {
        "UOB": UOB,
        "HLB": HLB,
        "MYB": MYB,
        "RHB": RHB,
        "PBB": PBB,
        "CIMB": CIMB,
    }

    
    
    
    def __init__(self, bank: str, excel_manager: ExcelManager):
        self.bank = self.BANK_CLASSES[bank]()
        self.bank_name = bank
        self.excel_manager = ExcelManager
        self.card_global_order = self._generate_ordered_list()
        self.card_column_map = {
            suffix: col for suffix, col, _ in self.card_global_order
        }
    

    def _generate_ordered_list(self) -> List[tuple]:
        return [
            (suffix, col, bank)
            for bank, cards in self.excel_manger.CARD_ORDERED_MAP.items()
            for suffix, col in cards
        ]


    def insert_new_card_column(self, new_card: str, after_card: str, excel_path: str):
        # Load the workbook and worksheet
        wb = load_workbook(excel_path)
        ws = wb.active

        # Find the index of the after_card in the global order
        idx = next((i for i, (suffix, _, _) in enumerate(self.card_global_order) if suffix == after_card), None)
        if idx is None:
            raise ValueError(f"After card {after_card} not found in global order.")

        # Get the column letter and index of the after_card
        after_col_letter = self.card_global_order[idx][1]
        after_col_index = column_index_from_string(after_col_letter)
        print(after_col_letter)

        # Insert a new column in the Excel sheet
        self.insert_column_with_format_and_merge(ws, after_col_index + 1)
        
        print("inserted new column")

        # Save the updated Excel file
        wb.save(excel_path)
        print(f"✔ Inserted new column for card {new_card} after {after_card} in Excel.")
        
    
    
        # Calculate the new column letter
        new_col_index = after_col_index + 1
        new_col_letter = get_column_letter(new_col_index)

        # Update the global order with the new card
        self.card_global_order.insert(idx + 1, (new_card, new_col_letter, self.bank_name))

        # Update the column letters for all subsequent cards in the global order
        for i in range(idx + 2, len(self.card_global_order)):
            old_suffix, old_col, old_bank = self.card_global_order[i]
            old_index = column_index_from_string(old_col)
            self.card_global_order[i] = (old_suffix, get_column_letter(old_index + 1), old_bank)
            print(self.card_global_order[i])  #debug

        # Update the card column map
        self.card_column_map = {
            suffix: col for suffix, col, _ in self.card_global_order
        }
        

        print(f"Inserted new column: {new_card} → {new_col_letter}")
        print("Updated card_column_map:")
        for k, v in self.card_column_map.items():
            print(f"  {k}: {v}")
        
        from collections import defaultdict

        new_card_map = defaultdict(list)
        for suffix, col, bank in self.card_global_order:
            new_card_map[bank].append([suffix, col])

        # Replace and save
        self.CARD_ORDERED_MAP = dict(new_card_map)

        self.save_card_order_map()   


    def parse_statement(self, pdf_path: str, password: Optional[str] = None) -> Dict[str, Dict[str, float]]:
        """Always generates both raw text and blocks files"""
        lines = TextExtractor(pdf_path, password)  # extract_text now saves raw text
        blocks = self.bank.create_blocks(lines)
        
        TextExtractor.save_blocks(pdf_path,blocks)
        
        if hasattr(self.bank, "extract"):
            
            return self.bank.extract(lines)
        else:
            
            return {
        
                card: self.bank.process_block(block, lines)
                for card, block in blocks.items()
            }   
        
    def write_to_excel(self, results: Dict[str, Dict[str, float]], excel_path: str, record_number: int):
        try:
            wb = load_workbook(excel_path)
            ws = wb.active

            base_row = 4 + (record_number - 1) * 7  # Row 4 is start of record 1
            bank_cards = dict(self.excel_manager.CARD_ORDERED_MAP[self.bank_name])
            print(bank_cards)
            for card, data in results.items():
                card_col = bank_cards.get(card)
                if not card_col:
                    print(f"Card {card} not found. Adding to sheet.")
                    bank_cards = [sfx for sfx, _, bk in self.card_global_order if bk == self.bank_name]
                    if bank_cards:
                        self.insert_new_card_column(card, bank_cards[-1], excel_path)
                        wb = load_workbook(excel_path)
                        ws = wb.active
                        bank_cards = dict(self.excel_manager.CARD_ORDERED_MAP[self.bank_name])  # update after insert
                        card_col = bank_cards.get(card)
                    else:
                        raise ValueError(f"No base card found for bank {self.bank_name} to insert {card}.")
                    ws[f"{card_col}{3}"] = card

                # Fill each value in the right row
                ws[f"{card_col}{base_row + 0}"] = data["previous_balance"]
                ws[f"{card_col}{base_row + 1}"] = data["credit_payment"]
                ws[f"{card_col}{base_row + 2}"] = data["debit_fees"]
                ws[f"{card_col}{base_row + 3}"] = data["retail_purchases"]
                ws[f"{card_col}{base_row + 4}"] = data["balance_due"]
                ws[f"{card_col}{base_row + 5}"] = data["minimum_payment"]

                print(f"✔ Updated card {card} in column {card_col}")

            wb.save(excel_path)
            print(f"✅ Excel updated and saved: {excel_path}")

        except Exception as e:
            raise RuntimeError(f"Failed to write to Excel: {str(e)}")
        
    

        
    