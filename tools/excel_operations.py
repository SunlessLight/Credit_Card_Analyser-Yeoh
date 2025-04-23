from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy
from openpyxl.worksheet.worksheet import Worksheet
import json

class ExcelManager:
    def __init__(self, card_ordered_map_path: Path):
        self.card_ordered_map_path = card_ordered_map_path
        self.CARD_ORDERED_MAP = self._load_card_order_map()
        
    def _load_card_order_map(self) -> dict:
        if not self.card_ordered_map_path.exists():
            with open(self.card_ordered_map_path, "w", encoding="utf-8") as f:
                json.dump({}, f, indent=2)
            return {}
        
        with open(self.card_ordered_map_path, "r", encoding="utf-8") as f:
            return json.load(f)
    
    def save_card_order_map(self, card_order_map: dict):
        try:
            with open(self.card_ordered_map_path, "w", encoding="utf-8") as f:
                json.dump(card_order_map, f, indent=2)
            print(f"✅ Saved card order map to {self.card_ordered_map_path}")
        except Exception as e:
            print(f"Failed to save card order map: {e}")

    def insert_column_with_format_and_merge(self, ws: Worksheet, idx: int):
        ws.insert_cols(idx)

        max_row = ws.max_row

        # Preserve styles and formatting
        for row in range(1, max_row + 1):
            old_cell = ws.cell(row=row, column=idx + 1)
            new_cell = ws.cell(row=row, column=idx)
            if old_cell.has_style:
                new_cell._style = copy(old_cell._style)
            new_cell.font = copy(old_cell.font)
            new_cell.border = copy(old_cell.border)
            new_cell.fill = copy(old_cell.fill)
            new_cell.number_format = copy(old_cell.number_format)
            new_cell.protection = copy(old_cell.protection)
            new_cell.alignment = copy(old_cell.alignment)

        # Copy column width
        col_letter_old = get_column_letter(idx + 1)
        col_letter_new = get_column_letter(idx)
        if col_letter_old in ws.column_dimensions:
            ws.column_dimensions[col_letter_new].width = ws.column_dimensions[col_letter_old].width

        # Handle merged cells
        updated_merges = []
        to_remove = []

        for merged in ws.merged_cells.ranges:
            min_col, max_col = merged.min_col, merged.max_col

            if max_col < idx:
                updated_merges.append(merged)
            elif min_col > idx:
                merged.shift(0, 1)
                updated_merges.append(merged)
            else:
                to_remove.append(merged)

        for r in to_remove:
            ws.unmerge_cells(range_string=str(r))

        ws.merged_cells.ranges = []
        for r in updated_merges:
            ws.merged_cells.add(r)